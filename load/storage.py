"""
load/storage.py
Persistência CSV/Excel para NF-e e NFS-e, lock de sessão, sincronização.
"""

import atexit, csv, os, shutil, time
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config.settings import (
    CABECALHO_CSV, CABECALHO_NFSE,
    CSV_PRINCIPAL, CSV_TEMP,
    EXCEL_PRINCIPAL, EXCEL_TEMP,
    CSV_NFSE_TEMP, EXCEL_NFSE_TEMP,
    CSV_NFSE_PRINCIPAL, EXCEL_NFSE_PRINCIPAL,
    LOCK_FILE, LOCK_TTL_SECONDS,
    LOG_TEMP, MODO_SESSAO, SESSAO_ID, TEMP_DIR, TEMP_TTL_SECONDS, USUARIO_ID,
)

# ── Lock / Sessão ──────────────────────────────────────────────────────────────

def criar_lock():
    try:
        with open(LOCK_FILE,"w") as f:
            f.write(f"Sessão: {SESSAO_ID}\nUsuário: {USUARIO_ID}\n"
                    f"Início: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
        return True
    except Exception:
        return False

def verificar_locks_ativos():
    try:
        meu = os.path.basename(LOCK_FILE)
        agora = time.time()
        return [n for n in os.listdir(TEMP_DIR)
                if n.startswith("lock_") and n != meu
                and (agora - os.path.getmtime(os.path.join(TEMP_DIR,n))) < LOCK_TTL_SECONDS]
    except Exception:
        return []

def _migrar_csv(caminho, cabecalho):
    """Garante que o CSV tenha exatamente as colunas do cabecalho atual.
    Se faltar colunas (versão antiga), reescreve o arquivo com as novas colunas vazias.
    Retorna True se precisou migrar, False se já estava correto."""
    if not os.path.exists(caminho) or os.path.getsize(caminho) == 0:
        return False
    try:
        for enc in ("utf-8","utf-8-sig","latin-1"):
            try:
                with open(caminho,"r",encoding=enc) as f:
                    rows = list(csv.DictReader(f))
                break
            except UnicodeDecodeError:
                continue
        else:
            return False

        # Verificar se cabeçalho já está correto
        if rows and set(cabecalho) == set(rows[0].keys()):
            return False  # já OK, sem migração

        # Migrar: reescrever com novo cabeçalho, preservando dados existentes
        with open(caminho,"w",newline="",encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=cabecalho, extrasaction="ignore")
            writer.writeheader()
            for row in rows:
                writer.writerow({k: row.get(k,"") for k in cabecalho})
        return True
    except Exception:
        return False

def inicializar_sessao():
    try:
        criar_lock()

        if MODO_SESSAO == "substituir":
            # Começa do zero — temp limpo, sem carregar o histórico
            _criar_csv_vazio(CSV_TEMP, CABECALHO_CSV)
            _criar_csv_vazio(CSV_NFSE_TEMP, CABECALHO_NFSE)
        else:
            # Modo acumular — carrega o histórico do principal para o temp
            if os.path.exists(CSV_PRINCIPAL):
                _migrar_csv(CSV_PRINCIPAL, CABECALHO_CSV)
                shutil.copy2(CSV_PRINCIPAL, CSV_TEMP)
            else:
                _criar_csv_vazio(CSV_TEMP, CABECALHO_CSV)

            if os.path.exists(CSV_NFSE_PRINCIPAL):
                _migrar_csv(CSV_NFSE_PRINCIPAL, CABECALHO_NFSE)
                shutil.copy2(CSV_NFSE_PRINCIPAL, CSV_NFSE_TEMP)
            else:
                _criar_csv_vazio(CSV_NFSE_TEMP, CABECALHO_NFSE)

        sincronizar_excel_temp()

        with open(LOG_TEMP,"w",encoding="utf-8") as f:
            f.write(f"Sessão: {SESSAO_ID}\nUsuário: {USUARIO_ID}\n"
                    f"Início: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n"
                    f"Modo: {MODO_SESSAO}\n")
        return True
    except Exception as e:
        print(f"Erro ao inicializar sessão: {e}")
        return False

# ── CSV ────────────────────────────────────────────────────────────────────────

def _criar_csv_vazio(caminho, cabecalho):
    with open(caminho,"w",newline="",encoding="utf-8") as f:
        csv.writer(f).writerow(cabecalho)

def salvar_produtos_csv(produtos, caminho=CSV_TEMP, cabecalho=None):
    if not produtos: return True, "Nenhum produto para salvar"
    if cabecalho is None: cabecalho = CABECALHO_CSV
    try:
        existe = os.path.exists(caminho) and os.path.getsize(caminho) > 0
        with open(caminho,"a",newline="",encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=cabecalho, extrasaction="ignore")
            if not existe:
                writer.writeheader()
            for p in produtos:
                writer.writerow({k: p.get(k,"") for k in cabecalho})
        return True, f"{len(produtos)} registro(s) salvos"
    except Exception as e:
        return False, f"Erro ao salvar CSV: {e}"

def salvar_nfse_csv(registros):
    return salvar_produtos_csv(registros, CSV_NFSE_TEMP, CABECALHO_NFSE)

def total_registros(caminho=CSV_TEMP):
    try:
        if not os.path.exists(caminho): return 0
        for enc in ("utf-8","utf-8-sig","latin-1"):
            try:
                with open(caminho,"r",encoding=enc) as f:
                    return max(0, sum(1 for _ in f) - 1)
            except UnicodeDecodeError:
                continue
        return 0
    except Exception:
        return 0

def carregar_chaves_nfse():
    chaves = set()
    if not os.path.exists(CSV_NFSE_TEMP): return chaves
    for enc in ("utf-8","utf-8-sig","latin-1"):
        try:
            with open(CSV_NFSE_TEMP,"r",encoding=enc) as f:
                for row in csv.DictReader(f):
                    chaves.add(f"{row.get('Chave_NFSe','')}_{row.get('Numero_NFSe','')}")
            return chaves
        except UnicodeDecodeError:
            continue
        except Exception:
            break
    return chaves

# ── Excel formatado ────────────────────────────────────────────────────────────

def _aplicar_formatacao_excel(caminho, sheet_name, titulo):
    """Aplica cabeçalho colorido, auto-largura e freeze no Excel."""
    try:
        wb = load_workbook(caminho)
        ws = wb[sheet_name]

        # Insere linha de título no topo
        ws.insert_rows(1)
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=ws.max_column)
        cell_titulo = ws.cell(row=1, column=1)
        cell_titulo.value = titulo
        cell_titulo.font      = Font(name="Segoe UI", size=12, bold=True, color="FFFFFF")
        cell_titulo.fill      = PatternFill("solid", fgColor="1A5276")
        cell_titulo.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # Formata cabeçalho (agora linha 2)
        fill_hdr = PatternFill("solid", fgColor="1F618D")
        font_hdr = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
        borda = Border(
            bottom=Side(style="thin", color="FFFFFF"),
            right= Side(style="thin", color="FFFFFF"),
        )
        for cell in ws[2]:
            cell.fill      = fill_hdr
            cell.font      = font_hdr
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            cell.border    = borda
        ws.row_dimensions[2].height = 32

        # Linhas alternadas
        fill_par  = PatternFill("solid", fgColor="EAF2FB")
        fill_imp  = PatternFill("solid", fgColor="FFFFFF")
        font_data = Font(name="Segoe UI", size=9)
        for row_idx in range(3, ws.max_row + 1):
            fill = fill_par if row_idx % 2 == 0 else fill_imp
            for cell in ws[row_idx]:
                cell.fill      = fill
                cell.font      = font_data
                cell.alignment = Alignment(vertical="center")

        # Auto-largura (máx 50)
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for row_idx in range(2, min(ws.max_row + 1, 200)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 50)

        # Freeze abaixo do cabeçalho
        ws.freeze_panes = "A3"

        wb.save(caminho)
    except Exception as e:
        print(f"Aviso formatação Excel: {e}")

def _df_para_excel(df, caminho, sheet_name, titulo):
    with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    _aplicar_formatacao_excel(caminho, sheet_name, titulo)

def _csv_para_df(caminho, cabecalho):
    for enc in ("utf-8", "utf-8-sig", "latin-1"):
        try:
            df = pd.read_csv(caminho, dtype=str, encoding=enc, on_bad_lines="skip")
            return df.reindex(columns=cabecalho, fill_value="")
        except UnicodeDecodeError:
            continue
        except Exception:
            break
    return pd.DataFrame(columns=cabecalho)

def sincronizar_excel_temp():
    if not os.path.exists(CSV_TEMP): return False
    try:
        df = _csv_para_df(CSV_TEMP, CABECALHO_CSV)
        _df_para_excel(df, EXCEL_TEMP, "Produtos_NFe",
                       "GCON/SIAN — NF-e — Produtos e Impostos")
        return True
    except Exception as e:
        print(f"Erro sincronizar Excel NF-e: {e}")
        return False

def sincronizar_excel_nfse_temp():
    if not os.path.exists(CSV_NFSE_TEMP): return False
    try:
        df = _csv_para_df(CSV_NFSE_TEMP, CABECALHO_NFSE)
        _df_para_excel(df, EXCEL_NFSE_TEMP, "Servicos_NFSe",
                       "GCON/SIAN — NFS-e — Notas de Serviço")
        return True
    except Exception as e:
        print(f"Erro sincronizar Excel NFS-e: {e}")
        return False

def atualizar_excel_principal():
    # Garante que o CSV principal existe antes de gerar Excel
    if not os.path.exists(CSV_PRINCIPAL):
        _criar_csv_vazio(CSV_PRINCIPAL, CABECALHO_CSV)
    try:
        df = _csv_para_df(CSV_PRINCIPAL, CABECALHO_CSV)
        _df_para_excel(df, EXCEL_PRINCIPAL, "Produtos_NFe",
                       "GCON/SIAN — NF-e — Produtos e Impostos")
        return True, f"Excel NF-e atualizado ({len(df)} registros)"
    except Exception as e:
        return False, f"Erro Excel NF-e: {e}"

def atualizar_excel_nfse_principal():
    # Garante que o CSV principal existe antes de gerar Excel
    if not os.path.exists(CSV_NFSE_PRINCIPAL):
        _criar_csv_vazio(CSV_NFSE_PRINCIPAL, CABECALHO_NFSE)
    try:
        df = _csv_para_df(CSV_NFSE_PRINCIPAL, CABECALHO_NFSE)
        _df_para_excel(df, EXCEL_NFSE_PRINCIPAL, "Servicos_NFSe",
                       "GCON/SIAN — NFS-e — Notas de Serviço")
        return True, f"Excel NFS-e atualizado ({len(df)} registros)"
    except Exception as e:
        return False, f"Erro Excel NFS-e: {e}"

def salvar_excel_sessao():
    """Após importar XMLs: gera Excel a partir do temp (sessão atual). Não toca no principal."""
    ok3, m3 = sincronizar_excel_temp()
    ok4, m4 = sincronizar_excel_nfse_temp()
    # Copia os Excels temporários para o local do principal (sobrescreve visualmente)
    resultados = {}
    for excel_temp, excel_principal, label in [
        (EXCEL_TEMP,      EXCEL_PRINCIPAL,      "excel_nfe"),
        (EXCEL_NFSE_TEMP, EXCEL_NFSE_PRINCIPAL, "excel_nfse"),
    ]:
        try:
            if os.path.exists(excel_temp):
                shutil.copy2(excel_temp, excel_principal)
                n = len(__import__('pandas').read_excel(excel_principal))
                resultados[label] = (True, f"Excel atualizado ({n} registros — sessão atual)")
            else:
                resultados[label] = (True, "Sem dados para gerar Excel")
        except Exception as e:
            resultados[label] = (False, f"Erro Excel: {e}")
    return resultados

def salvar_tudo():
    """Sincronizar Tudo: temp→principal (substitui) e regera Excel do principal."""
    ok1, m1 = sincronizar_com_principal()
    ok2, m2 = sincronizar_nfse_com_principal()
    ok3, m3 = atualizar_excel_principal()
    ok4, m4 = atualizar_excel_nfse_principal()
    return {
        "csv_nfe":   (ok1, m1),
        "csv_nfse":  (ok2, m2),
        "excel_nfe": (ok3, m3),
        "excel_nfse":(ok4, m4),
    }

# ── Sincronização temp → principal ─────────────────────────────────────────────

def _sincronizar_csv(csv_temp, csv_principal, cabecalho, chave_fn):
    try:
        if not os.path.exists(csv_temp):
            return True, "Nenhum dado temporário"

        novas = []
        for enc in ("utf-8","utf-8-sig","latin-1"):
            try:
                with open(csv_temp,"r",encoding=enc) as f:
                    novas = list(csv.DictReader(f))
                break
            except UnicodeDecodeError:
                continue
        if not novas:
            return True, "Nenhum dado novo"

        # Backup do principal antes de qualquer escrita
        if os.path.exists(csv_principal):
            nome_backup = os.path.basename(csv_principal).replace(".csv", f"_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")
            backup = os.path.join(TEMP_DIR, nome_backup)
            try: shutil.copy2(csv_principal, backup)
            except Exception: pass

        if MODO_SESSAO == "substituir":
            # Sobrescreve o principal com exatamente o que veio desta sessão
            with open(csv_principal,"w",newline="",encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=cabecalho, extrasaction="ignore")
                writer.writeheader()
                for row in novas:
                    writer.writerow({c: row.get(c,"") for c in cabecalho})
            return True, f"{len(novas)} registro(s) salvos (substituição total)"

        else:
            # Modo acumular — append deduplicado
            if not os.path.exists(csv_principal):
                _criar_csv_vazio(csv_principal, cabecalho)
            _migrar_csv(csv_principal, cabecalho)

            chaves = set()
            for enc in ("utf-8","utf-8-sig","latin-1"):
                try:
                    with open(csv_principal,"r",encoding=enc) as f:
                        for row in csv.DictReader(f):
                            chaves.add(chave_fn(row))
                    break
                except UnicodeDecodeError:
                    continue

            add = 0
            with open(csv_principal,"a",newline="",encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=cabecalho, extrasaction="ignore")
                for row in novas:
                    k = chave_fn(row)
                    if k not in chaves:
                        writer.writerow({c: row.get(c,"") for c in cabecalho})
                        chaves.add(k)
                        add += 1
            return True, f"{add} registro(s) sincronizados"

    except Exception as e:
        return False, f"Erro: {e}"

def sincronizar_com_principal():
    return _sincronizar_csv(
        CSV_TEMP, CSV_PRINCIPAL, CABECALHO_CSV,
        lambda r: f"{r.get('Chave_NFe','')}_{r.get('Item','')}_{r.get('cProd','')}",
    )

def sincronizar_nfse_com_principal():
    return _sincronizar_csv(
        CSV_NFSE_TEMP, CSV_NFSE_PRINCIPAL, CABECALHO_NFSE,
        lambda r: f"{r.get('Chave_NFSe','')}_{r.get('Numero_NFSe','')}",
    )

# ── Limpeza ────────────────────────────────────────────────────────────────────

def limpar_temporarios():
    for c in [CSV_TEMP, EXCEL_TEMP, LOG_TEMP, LOCK_FILE, CSV_NFSE_TEMP, EXCEL_NFSE_TEMP]:
        try:
            if os.path.exists(c): os.remove(c)
        except Exception:
            pass
    try:
        agora = time.time()
        for n in os.listdir(TEMP_DIR):
            # Limpa temporários expirados e backups com mais de 7 dias
            eh_temp   = any(n.startswith(p) for p in ("temp_","lock_"))
            eh_backup = "_backup_" in n and n.endswith(".csv")
            c = os.path.join(TEMP_DIR, n)
            if not os.path.isfile(c): continue
            idade = agora - os.path.getmtime(c)
            if eh_temp and idade > TEMP_TTL_SECONDS:
                try: os.remove(c)
                except Exception: pass
            elif eh_backup and idade > 7 * 86400:
                try: os.remove(c)
                except Exception: pass
    except Exception:
        pass

atexit.register(limpar_temporarios)
